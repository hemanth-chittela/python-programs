import pandas as pd
import openpyxl as cat
import re

def merging_sheets(path):
    sheet=cat.load_workbook(path)
    sheet1=sheet["cat"]
    sheet2=sheet["cat_fix"]
    max_row=sheet2.max_row
    max_column=sheet2.max_column
    next_row=sheet1.max_row + 1
    print(max_row,max_column)
    for row_in_sheet in range(2, max_row + 1):  # Use range(1, max_row + 1) to include all rows
        for col_in_sheet in range(1, max_column + 1):  # Use range(1, max_column + 1) to include all columns
            appending_rows = sheet2.cell(row=row_in_sheet, column=col_in_sheet).value
            sheet1.cell(row=next_row, column=col_in_sheet).value = appending_rows
        next_row+=1
    sheet.save(path)
    print("successfully appended")
    sheet.close()
merging_sheets('C:/Users/pc/Desktop/Sample.xlsx')
