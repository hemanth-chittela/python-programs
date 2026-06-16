import sys
import pandas as pd
from openpyxl import load_workbook

# Get the path to summary.xlsx from command-line arguments
summary_file_path = str(sys.argv[1]) + "/Summary.xlsx"

# Load the existing Excel file
workbook = load_workbook(summary_file_path)

# Check if 'Sheet1' exists and read it into a DataFrame
if 'Sheet1' in workbook.sheetnames:
    sheet1_df = pd.read_excel(summary_file_path, sheet_name='Sheet1')
    
    # Count non-empty rows in 'Sheet1'
    # Check all rows for non-NaN values in all columns
    filled_data_count = sheet1_df.dropna(how='all').shape[0]
else:
    filled_data_count = 0  # If 'Sheet1' does not exist, set count to 0

# Create the data for 'Summary1' sheet
summary1_data = {
    'fruits': ['sheet1'],
    'count': [filled_data_count]
}

summary1_df = pd.DataFrame(summary1_data)

# Load the workbook and add the new sheet with the data
with pd.ExcelWriter(summary_file_path, engine='openpyxl', mode='a') as writer:
    summary1_df.to_excel(writer, sheet_name='Summary1', index=False)

print(f"'Summary1' sheet created in {summary_file_path} with the count of filled data from 'Sheet1'.")
