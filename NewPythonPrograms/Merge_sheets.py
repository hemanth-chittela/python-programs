import pandas as pd
import openpyxl as cat
import re

def merging_sheets(path):
    sheet=cat.load_workbook(path)
    sheet1=sheet["cat"]
    sheet_names=sheet.sheetnames
    next_row = sheet1.max_row + 1
    for name in sheet_names:
        if re.search("[^cat_]",name):
            print(name)
            current_sheet=sheet[name]
            max_row=current_sheet.max_row
            max_column=current_sheet.max_column
            print(max_row,max_column)
            for row_in_sheet in range(2, max_row + 1):  # Use range(1, max_row + 1) to include all rows
                for col_in_sheet in range(1, max_column + 1):  # Use range(1, max_column + 1) to include all columns
                    appending_rows = current_sheet.cell(row=row_in_sheet, column=col_in_sheet).value
                    sheet1.cell(row=next_row, column=col_in_sheet).value = appending_rows
                next_row+=1
    sheet.save(path)
    print("Successfully appended")
    sheet.close()
merging_sheets('C:/Users/pc/Desktop/Sample.xlsx')

